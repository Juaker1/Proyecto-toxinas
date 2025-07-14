import pandas as pd
import io
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def generate_excel(data_dict, filename_prefix, sheet_names=None, metadata=None):
    """
    Generate Excel file from dictionary of dataframes or a single dataframe.
    
    Parameters:
    - data_dict: Dict of dataframes or single dataframe
    - filename_prefix: Prefix for the filename
    - sheet_names: Custom sheet names (optional)
    - metadata: Dictionary of metadata to include in a separate sheet
    
    Returns:
    - Tuple of (io.BytesIO, filename)
    """
    output = io.BytesIO()
    
    # Get timestamp for filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{filename_prefix}_{timestamp}.xlsx"
    
    # Check if data_dict is a single dataframe
    if isinstance(data_dict, pd.DataFrame):
        data_dict = {"Data": data_dict}
    
    # If sheet_names is provided, use it
    if sheet_names and len(sheet_names) == len(data_dict):
        sheet_dict = {name: df for name, df in zip(sheet_names, data_dict.values())}
    else:
        sheet_dict = data_dict
    
    # Create Excel writer
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # First add a metadata sheet if provided
        if metadata:
            # Convert metadata dictionary to a DataFrame for display
            meta_items = [[k, v] for k, v in metadata.items()]
            meta_df = pd.DataFrame(meta_items, columns=['Propiedad', 'Valor'])
            meta_df.to_excel(writer, sheet_name='Metadatos', index=False)
            
            # Format the metadata sheet
            worksheet = writer.sheets['Metadatos']
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col_num, value in enumerate(meta_df.columns.values, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                
            # Auto-adjust column widths
            for col_num, column in enumerate(meta_df.columns, 1):
                worksheet.column_dimensions[get_column_letter(col_num)].width = 25
        
        # Then add all data sheets
        for sheet_name, df in sheet_dict.items():
            # Clean sheet name (Excel has 31 char limit and no special chars)
            clean_sheet_name = str(sheet_name).replace('/', '_').replace('\\', '_')
            clean_sheet_name = ''.join(c for c in clean_sheet_name if c.isalnum() or c in ['_', '-'])
            clean_sheet_name = clean_sheet_name[:31]  # Excel sheet name length limit
            
            # Write dataframe to Excel
            df.to_excel(writer, sheet_name=clean_sheet_name, index=False)
            
            # Apply formatting
            worksheet = writer.sheets[clean_sheet_name]
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            centered_alignment = Alignment(horizontal='center', vertical='center')
            
            # Format headers
            for col_num, value in enumerate(df.columns.values, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = centered_alignment
                
            # Auto-adjust column widths
            for col_num, column in enumerate(df.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_num)
                
                # Consider header width
                header_length = len(str(column)) + 2
                
                # Review cell contents
                for cell_value in df[column]:
                    try:
                        if len(str(cell_value)) > max_length:
                            max_length = len(str(cell_value))
                    except:
                        pass
                
                adjusted_width = max(max_length + 2, header_length)
                adjusted_width = min(adjusted_width, 40)  # Maximum width of 40 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    return output, filename